"""
LifeOps API — Data Exporter Service (CSV & Excel .xlsx)
========================================================
Exports raw user data into clean, portable CSV and multi-sheet
Excel (.xlsx) files with UTF-8 BOM encoding for seamless Excel opening.
"""
import io
import csv
import datetime as dt
from typing import Dict, Any, List

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from services.supabase_client import db


# ── Entity Fetchers ───────────────────────────────

def _fetch_sport_data(user_id: str) -> List[Dict[str, Any]]:
    query = (
        db("activities")
        .select("*, workouts(*)")
        .eq("user_id", user_id)
        .eq("activity_type", "sport")
        .order("date", desc=True)
    )
    raw = query.execute().data or []
    rows = []
    for r in raw:
        w = r.get("workouts", {})
        if isinstance(w, list) and w:
            w = w[0]
        elif not isinstance(w, dict):
            w = {}
        rows.append({
            "id": r.get("id"),
            "fecha": r.get("date"),
            "titulo": r.get("title"),
            "tipo_deporte": w.get("workout_type", "sport"),
            "distancia_km": w.get("distance_km") or "",
            "duracion_min": r.get("duration_minutes") or "",
            "calorias": w.get("calories") or "",
            "frecuencia_cardiaca_avg": w.get("heart_rate_avg") or "",
            "record_personal": "Sí" if w.get("personal_best") else "No",
            "notas": r.get("notes") or "",
        })
    return rows


def _fetch_books_data(user_id: str) -> List[Dict[str, Any]]:
    query = (
        db("activities")
        .select("*, books(*)")
        .eq("user_id", user_id)
        .eq("activity_type", "book")
        .order("date", desc=True)
    )
    raw = query.execute().data or []
    rows = []
    for r in raw:
        b = r.get("books", {})
        if isinstance(b, list) and b:
            b = b[0]
        elif not isinstance(b, dict):
            b = {}
        
        pages_read = b.get("pages_read") or 0
        pages_total = b.get("pages_total") or 0
        pct = round((pages_read / pages_total * 100), 1) if pages_total > 0 else 0

        rows.append({
            "id": r.get("id"),
            "fecha_registro": r.get("date"),
            "titulo": r.get("title"),
            "autor": b.get("author") or "",
            "genero": b.get("genre") or "",
            "paginas_leidas": pages_read,
            "paginas_totales": pages_total,
            "progreso_pct": f"{pct}%",
            "estado": b.get("status") or "reading",
            "valoracion_estrellas": r.get("rating") or "",
            "notas": r.get("notes") or "",
        })
    return rows


def _fetch_films_data(user_id: str) -> List[Dict[str, Any]]:
    query = (
        db("activities")
        .select("*, films(*)")
        .eq("user_id", user_id)
        .eq("activity_type", "film")
        .order("date", desc=True)
    )
    raw = query.execute().data or []
    rows = []
    for r in raw:
        f = r.get("films", {})
        if isinstance(f, list) and f:
            f = f[0]
        elif not isinstance(f, dict):
            f = {}

        rows.append({
            "id": r.get("id"),
            "fecha_visto": r.get("date"),
            "titulo": r.get("title"),
            "tipo_medio": f.get("media_type") or "film",
            "genero": f.get("genre") or "",
            "año": f.get("year") or "",
            "plataforma": f.get("platform") or "",
            "director": f.get("director") or "",
            "valoracion_estrellas": r.get("rating") or "",
            "reseña": f.get("review") or r.get("notes") or "",
        })
    return rows


def _fetch_tasks_data(user_id: str) -> List[Dict[str, Any]]:
    query = (
        db("tasks")
        .select("*")
        .eq("user_id", user_id)
        .order("due_date", desc=False)
    )
    raw = query.execute().data or []
    rows = []
    for t in raw:
        rows.append({
            "id": t.get("id"),
            "titulo": t.get("title"),
            "id_proyecto": t.get("project_id") or "",
            "estado": t.get("status") or "todo",
            "prioridad": t.get("priority") or "medium",
            "horas_estimadas": t.get("estimated_hours") or "",
            "horas_reales": t.get("actual_hours") or "",
            "fecha_limite": t.get("due_date") or "",
            "fecha_completada": t.get("completed_at") or "",
            "descripcion": t.get("description") or "",
        })
    return rows


def _fetch_projects_data(user_id: str) -> List[Dict[str, Any]]:
    query = (
        db("projects")
        .select("*")
        .eq("user_id", user_id)
        .order("created_at", desc=True)
    )
    raw = query.execute().data or []
    rows = []
    for p in raw:
        rows.append({
            "id": p.get("id"),
            "nombre": p.get("name"),
            "estado": p.get("status") or "active",
            "prioridad": p.get("priority") or "medium",
            "presupuesto_eur": p.get("budget") or "",
            "horas_estimadas": p.get("estimated_hours") or "",
            "horas_reales": p.get("actual_hours") or "",
            "fecha_inicio": p.get("start_date") or "",
            "fecha_objetivo": p.get("target_end_date") or "",
            "descripcion": p.get("description") or "",
        })
    return rows


# ── CSV Generator ─────────────────────────────────

ENTITY_FETCHERS = {
    "sport": ("Entrenamientos", _fetch_sport_data),
    "books": ("Biblioteca_Libros", _fetch_books_data),
    "films": ("Cine_Series", _fetch_films_data),
    "tasks": ("Tareas_Kanban", _fetch_tasks_data),
    "projects": ("Proyectos", _fetch_projects_data),
}


def export_entity_csv(entity: str, user_id: str) -> bytes:
    """
    Generates a UTF-8 BOM encoded CSV bytes buffer for a specific entity.
    """
    if entity not in ENTITY_FETCHERS:
        raise ValueError(f"Entidad no soportada: {entity}")

    _, fetcher = ENTITY_FETCHERS[entity]
    data = fetcher(user_id)

    output = io.StringIO()
    # Write UTF-8 BOM so Excel opens accents without mojibake
    output.write("\ufeff")

    if data:
        fieldnames = list(data[0].keys())
        writer = csv.DictWriter(output, fieldnames=fieldnames, delimiter=";", quoting=csv.QUOTE_MINIMAL)
        writer.writeheader()
        for row in data:
            writer.writerow(row)
    else:
        output.write("No hay registros disponibles en esta sección\n")

    return output.getvalue().encode("utf-8")


# ── Excel (.xlsx) Multi-Sheet Generator ───────────

def export_full_excel(user_id: str) -> io.BytesIO:
    """
    Generates a styled multi-sheet Excel (.xlsx) workbook containing all user data.
    """
    wb = Workbook()
    # Remove default sheet
    wb.remove(wb.active)

    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    sheets_config = [
        ("🏃 Deporte & Fitness", _fetch_sport_data),
        ("📚 Biblioteca de Libros", _fetch_books_data),
        ("🎬 Cine & Series", _fetch_films_data),
        ("📋 Tablero de Tareas", _fetch_tasks_data),
        ("💼 Portafolio Proyectos", _fetch_projects_data),
    ]

    for title, fetcher in sheets_config:
        ws = wb.create_sheet(title=title)
        ws.views.sheetView[0].showGridLines = True

        data = fetcher(user_id)
        if data:
            headers = list(data[0].keys())
            # Format header labels
            display_headers = [h.replace("_", " ").upper() for h in headers]
            ws.append(display_headers)

            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")

            ws.row_dimensions[1].height = 24

            # Write rows
            for row_idx, item in enumerate(data, start=2):
                row_vals = [item[k] for k in headers]
                ws.append(row_vals)
                for col_idx in range(1, len(headers) + 1):
                    c = ws.cell(row=row_idx, column=col_idx)
                    c.font = data_font
                    c.border = thin_border

            # Auto-adjust column widths
            for col in ws.columns:
                max_len = max(len(str(cell.value or "")) for cell in col)
                col_letter = get_column_letter(col[0].column)
                ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
        else:
            ws.append(["Sin datos registrados en esta sección"])
            ws.cell(row=1, column=1).font = data_font

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
